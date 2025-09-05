import os
import openpyxl

class ExcelReader:
    def __init__(self, file_path=None):
        if not file_path:
            file_path = os.path.join(os.path.dirname(__file__), '../Resources/Data.xlsx')
        self.workbook = openpyxl.load_workbook(file_path)
        self.sheet = None

    def select_sheet(self, sheet_name=None, index=None):
        """Select a sheet by name or index"""
        if sheet_name:
            if sheet_name in self.workbook.sheetnames:
                self.sheet = self.workbook[sheet_name]
            else:
                raise ValueError(f"Sheet '{sheet_name}' not found")
        elif index is not None:
            sheets = self.workbook.sheetnames
            if 0 <= index < len(sheets):
                self.sheet = self.workbook[sheets[index]]
            else:
                raise IndexError(f"Sheet index {index} out of range")
        else:
            self.sheet = self.workbook.active
        return self.sheet

    def get_cell_value(self, row, column):
        return self.sheet.cell(row=row, column=column).value

    def get_row_as_dict(self, row):
        """Return row values mapped to headers in first row"""
        headers = [cell.value for cell in self.sheet[1]]
        values = [cell.value for cell in self.sheet[row]]
        return dict(zip(headers, values))
